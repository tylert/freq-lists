#!/usr/bin/env python


#   https://openpyxl.readthedocs.io/en/stable/index.html
#   https://realpython.com/openpyxl-excel-spreadsheets-python/#writing-excel-spreadsheets-with-openpyxl
#   https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size


from csv import reader

import click
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


def drop_a_deuce(csv_filename=None, xlsx_filename=None):
    ''' '''
    # Create a new spreadsheet and set it all up
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Channels'
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER

    # Each row from the CSV file becomes a row in the spreadsheet
    with open(csv_filename, 'r') as csv_file:
        rows = reader(csv_file)
        for row in rows:
            sheet.append(row)

    # Set the font and style of each cell
    for row in sheet.rows:
        for cell in row:
            cell.font = Font(name='Quicksand', size=8)
            cell.border = Border(bottom=Side(border_style='thin', color='000000'))

    # Set the column widths
    # Autofit/bestfit doesn't work
    # This is a mega ugly hack to try to get the column widths to be sensible
    MIN_WIDTH = 10
    for i, column_cells in enumerate(sheet.columns, start=1):
        width = (
            length
            if (length := max(len(str(cell_value) if (cell_value := cell.value) is not None else '')
                              for cell in column_cells)) >= MIN_WIDTH
            else MIN_WIDTH
        )
        sheet.column_dimensions[get_column_letter(i)].width = width * 0.75

    # Write the xlsx file out
    workbook.save(filename=xlsx_filename)


@click.command()
@click.option(
    '--input_file',
    '-i',
    default=None,
    help='Input file',
)
@click.option(
    '--output_file',
    '-o',
    default=None,
    help='Output file',
)
def main(input_file, output_file):
    ''' '''
    drop_a_deuce(csv_filename=input_file, xlsx_filename=output_file)


if __name__ == '__main__':
    main()
